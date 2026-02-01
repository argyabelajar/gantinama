from openpyxl.utils import rows_from_range
from openpyxl.worksheet.cell_range import CellRange
from copy import copy
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from Color import Colors


def apply_border_to_merge(ws, cr, border):
    for row in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row,
                            min_col=cr.min_col, max_col=cr.max_col):
        for cell in row:
            cell.border = border


def is_cell_styled(cell):
    if cell.value is not None:
        return True
    
    fill = cell.fill

    if isinstance(fill, PatternFill) and cell.fill and cell.fill.patternType is not None and cell.fill.patternType != 'none':
        return True
    if cell.border and any([
        cell.border.left and cell.border.left.style,
        cell.border.right and cell.border.right.style,
        cell.border.top and cell.border.top.style,
        cell.border.bottom and cell.border.bottom.style
        ]):
        return True
    return False
    
def get_merge_info(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return merged_range
    return None

def copy_range(range_str, sheet, target_range):
    min_row = CellRange(range_str).min_row
    for row in rows_from_range(range_str):
        for cell in row:
            if (sheet[cell].value is not None or is_cell_styled(sheet[cell])) and not isinstance(sheet[cell], MergedCell):

                dst_cell = sheet[cell].offset(row=(target_range.min_row - min_row), column=0)
                src_cell = sheet[cell]

                dst_cell.value = src_cell.value

                dst_cell.font = copy(src_cell.font)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)

                dst_cell.number_format = src_cell.number_format


def get_merge_list(ws, r_range , r_offset):
    area = CellRange(r_range)  
    mlist: list[CellRange] = []  
    for mc in ws.merged_cells:
#        print(mc.coord, area)
        if mc.coord not in area:
            continue
        cr = CellRange(mc.coord)
        cr.shift(row_shift=r_offset)
        if mc.max_row > mc.min_row:
            cr.shift(row_shift=(mc.max_row-mc.min_row))
#        print(cr.coord)
        mlist.append(cr)

    return mlist

def get_empty_row_length(ws, target_range_coor) -> int:
    

    sum_empty_space:int = 0
#    print(target_range_coor.min_row, ws.max_row + 1)
    for i in range(target_range_coor.min_row, ws.max_row + 1):
        target_cell = ws.cell(row = i, column=1)
        if is_cell_styled(target_cell) or get_merge_info(ws, row=i, col=1):
            break
        else:
            sum_empty_space += 1
    
    return sum_empty_space

def insert_rows_with_merges(ws, idx, amount=1):


    merges_to_update = []
    merges_to_keep = []
    
    for merged_range in list(ws.merged_cells.ranges):
        cr = CellRange(merged_range.coord)
        
        if cr.min_row >= idx:
            merges_to_update.append(cr)
        else:
            merges_to_keep.append(cr)
    

    for cr in merges_to_update:
        ws.unmerge_cells(cr.coord)
    
    ws.insert_rows(idx, amount)
    
    for cr in merges_to_update:
        cr.shift(row_shift=amount)
        ws.merge_cells(cr.coord)


def copy_from_another_row(ws, row_range, row_offset, border = None):
    range_coor = CellRange(row_range)
    target_range_coor = range_coor
    target_range_coor.shift(row_shift=row_offset)

#    print(target_range_coor.coord)
    
    new_merge_list: list[CellRange] = get_merge_list(ws, row_range, row_offset)
    biggest_row_size = 0
    biggest_merged_size = 0

    for merged in new_merge_list:
#        print(merged.coord)
        size = merged.max_row - merged.min_row
        if size > biggest_merged_size:
            biggest_merged_size = size

    target_range_coor.shift(row_shift=biggest_merged_size)
#    print(target_range_coor.coord)

    for coord in new_merge_list:
        size = target_range_coor.max_row - target_range_coor.min_row + 1
        if size > biggest_row_size:
            biggest_row_size = size

    empty_row_size = get_empty_row_length(ws, target_range_coor)

    print(f"{Colors.MAGENTA}Row required status:{Colors.ENDC}",f" empty = {empty_row_size}, big = {biggest_row_size}")

    if biggest_row_size + 1 > empty_row_size: #tambah 1 menghitung spasi
        to_add = biggest_row_size - empty_row_size + 1 #tambah 1 untuk spasi
        place = target_range_coor.max_row
        print(f"{Colors.BLUE}New row to insert:{Colors.ENDC}",f" place = {place}, to add = {to_add}")
        insert_rows_with_merges(ws, place, to_add)


    for nm in new_merge_list:
        ws.merge_cells(nm.coord)


    copy_range(row_range, ws, target_range_coor)
    if border is not None:
        style_border = Border(
        left=Side(style=border),
        right=Side(style=border),
        top=Side(style=border),
        bottom=Side(style=border))

        apply_border_to_merge(ws, target_range_coor, style_border)

def get_style_row_range(ws, col, row):
    
    if col > 1:
        for i in range(col, 0, -1):
            cell = ws.cell(column=i, row=row)
            if cell.value is None and not is_cell_styled(cell) and not get_merge_info(ws, row, i):
                break
            lcol = i
    
    for i in range(col, ws.max_column + 1):
        cell = ws.cell(column=i, row=row)
        if cell.value is None and not is_cell_styled(cell) and not get_merge_info(ws, row, i):
            break
        rcol = i
    
    for i in range(row, ws.max_row + 1):
        cell = ws.cell(column=col, row=i)
        if cell.value is None and not is_cell_styled(cell) and not get_merge_info(ws, row, i):
            break
        mrow = i

    lcell = ws.cell(column=lcol, row=row)
    rcell = ws.cell(column=rcol, row=mrow)
    return f"{lcell.coordinate}:{rcell.coordinate}"
