
###Pengubah Nama XLXS

"""
Note:
- Program ini cuma untuk ubah nama, tambah row baru dengan style dari row sebelumnya,
  dan hapus row tidak berguna
- Nilai dari cell excel dihapus mandiri
- Spacing di akhir perlu diatur sendiri
- Kontak: email: argya.menuntut.ilmu@gmail.com, ig: @argya.rayyan
"""

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from copier import copy_from_another_row, get_style_row_range, is_cell_styled
from openpyxl.worksheet.cell_range import CellRange
from Color import Colors

column_for_name = int(input("Kolom yang berisi nama (dalam bilangan bulat; contoh: kalau kolom A tulis: 1, kalau B tulis: 2, dsb): "))
#Waktu penulisan program, semua cell nama ada di kolom B, jadi agar lebih efisien iterasinya ada di kolom tertentu
##Kalau namanya ada di beberapa kolom, coba jalankan beberapa kali
border = input("Jenis border (thin, bold, dsb)(kalau gak ada tulis: none): ").lower()
#Kalau ada border, tulis jenis bordernya
input_path = input("Nama input (contoh: goo.xlsx) (jangan lupa tambah .xlxs): ")
#Nama file yang mau diubah
output_path = input("Nama output (contoh: foo.xlsx) (jangan lupa tambah .xlxs): ")
#Nama file hasil
name_path = input("Nama file nama (jangan lupa .txt): ")

wb = load_workbook(input_path)

if border == "none":
    border = None

if 'Sheet1' in wb.sheetnames:
    del wb['Sheet1']

def isDivisi(string: str):
    divisi = ["ORKES", "BAHASA", "WU", "MEDKOM", "PSDM", "PDS", "PMS", "BPH"]
    return string.upper() in divisi

def is_in_merged_range(ws, cell):
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return True
    return False

def delete_rows_with_cleanup(ws, idx, amount=1):
    for row in range(idx, idx + amount):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue

            cell.value = None
            cell.style = 'Normal'


            from openpyxl.styles import Font, Fill, Border, Alignment, Protection
            cell.font = Font()
            cell.fill = Fill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.protection = Protection()
            cell.number_format = 'General'
    

    merges_to_remove = []
    merges_to_update = []
    merges_to_keep = []
    
    for merged_range in list(ws.merged_cells.ranges):
        cr = CellRange(merged_range.coord)
        
        if cr.max_row < idx:
            merges_to_keep.append(cr)
        elif cr.min_row >= idx + amount:
            merges_to_update.append(cr)
        else:
            merges_to_remove.append(cr)
    
    for cr in merges_to_remove + merges_to_update:
        ws.unmerge_cells(cr.coord)
    
    ws.delete_rows(idx, amount)
    
    for cr in merges_to_update:
        cr.shift(row_shift=-amount)
        ws.merge_cells(cr.coord)


tempname: list[str] = []
with open(name_path, "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line:
            continue

        if line.endswith(":"):
            tempname.append(line[:-1])
        else:
            tempname.append(line)

new_names: dict[list] = {}
currentKey = ""
for i in tempname:
    if isDivisi(i):
        new_names[i] = []
        currentKey = i
    else:
        new_names[currentKey].append(i)
    

for ws in wb.worksheets:
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=column_for_name)
        if isinstance(cell.value, str) and cell.value.strip().lower() == "nama":
            r = row + 1

            while True:
                if isinstance(ws.cell(row=r, column=column_for_name), MergedCell):
                    r += 1
                else:
                    break
            name_idx = 0
            while r <= ws.max_row and name_idx < len(new_names[ws.title.upper()]):

                target = ws.cell(row=r, column=column_for_name)
                if isinstance(target, MergedCell):
                    print(f"{Colors.GREEN}Cell status: {Colors.ENDC}{Colors.RED}MERGED{Colors.ENDC}")
                    r += 1
                    continue

                if target.value is None and ws.cell(row=r, column=1).value is None:
                    style_row = r-1
                    style_col = column_for_name

                    while isinstance(ws.cell(column=style_col, row=style_row), MergedCell):
                        style_row -= 1

                    style_range = get_style_row_range(ws=ws, row=style_row, col=style_col)
                    copy_from_another_row(ws, style_range, 1, border)

                print(f"{Colors.GREEN}Edited:{Colors.ENDC} {target.value}, {new_names[ws.title.upper()][name_idx]}, {r}")
                target.value = new_names[ws.title.upper()][name_idx]
                r += 1
                name_idx += 1
            
            
            if (name_idx >= len(new_names[ws.title.upper()])) and r <= ws.max_row:
                cur_row = r
                cur_col = column_for_name
                current_cell = ws.cell(row=cur_row, column=cur_col)

                while isinstance(current_cell, MergedCell):
                    cur_row += 1
                    r += 1
                    current_cell = ws.cell(row=cur_row, column=cur_col)

                while current_cell.value is not None or is_cell_styled(current_cell):
                    delete_rows_with_cleanup(ws, cur_row, 1)


wb.save(output_path)
